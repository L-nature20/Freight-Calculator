"""Excel 导入 / 导出服务"""
import os
import io
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from ..models import db, DeliveryOrder, LtlApproval, Contract, ContractRate


# ── 通用样式 ──
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── 交货单列定义 ──
DELIVERY_COLUMNS = [
    '交货单号', '装运点', '装运点描述', '运输区域', '运输区域描述',
    '总重量（吨）', '总体积（m³）', '拼车单号', '运单号', '发货过账日期',
    '车牌号', '销售组织', '组织名称', '承运商编码', '承运商名称',
    '送达方编码', '送达方名称', '货物类型', '运输方式',
]

DELIVERY_REQUIRED = [
    '交货单号', '装运点', '装运点描述', '运输区域', '运输区域描述',
    '总重量（吨）', '总体积（m³）', '发货过账日期',
    '销售组织', '组织名称', '承运商名称', '承运商编码',
    '送达方名称', '送达方编码',
]

DELIVERY_NUMERIC = {'总重量（吨）', '总体积（m³）'}

# ── 零担审批列定义 ──
LTL_COLUMNS = ['零担类型', '车序号', '交货单号', '零担审批月份']
LTL_REQUIRED = ['零担类型', '车序号', '交货单号', '零担审批月份']

# ── 合同列定义 ──
CONTRACT_COLUMNS = [
    '合同编码', '合同名称', '承运商名称', '承运商编码',
    '合同有效期起始', '合同有效期截止', '合同状态',
]
CONTRACT_REQUIRED = [
    '合同编码', '合同名称', '承运商名称', '承运商编码',
    '合同有效期起始', '合同有效期截止',
]

# ── 费率列定义 ──
RATE_COLUMNS = [
    '装运点', '装运点描述', '运输区域', '运输区域描述',
    '省/直辖市', '地级市', '县/区', '里程(KM)', '线路类型',
    '增值税税率', '含增值税运输价格1', '含增值税运输价格2',
    '含增值税运输价格3', '含增值税运输价格4', '含增值税末端费用', '货物类型',
    '运输方式',
]
RATE_REQUIRED = [
    '装运点', '装运点描述', '运输区域', '运输区域描述',
    '省/直辖市', '地级市', '县/区', '里程(KM)', '线路类型',
    '增值税税率', '运输方式',
]
RATE_NUMERIC = {'里程(KM)', '含增值税运输价格1', '含增值税运输价格2',
                '含增值税运输价格3', '含增值税运输价格4', '含增值税末端费用'}


# ─────────────────────────────────────────
#  交货单导入
# ─────────────────────────────────────────
def import_delivery_orders(file_stream, on_duplicate='skip'):
    """
    导入交货单Excel。
    全有或全无：数据质量错误则整批拒绝。
    on_duplicate: 'skip'=跳过重复  |  'overwrite'=覆盖
    数据库已有不算错误，按用户选择处理；批次内重复仍然报错。
    返回 (success_count, errors_list)
    """
    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    # 表头映射
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    col_map = {}
    for i, h in enumerate(header_row):
        if h and str(h).strip() in DELIVERY_COLUMNS:
            col_map[str(h).strip()] = i

    # 一次性加载已有交货单到 dict（delivery_no → existing对象）
    existing_map = {
        r.delivery_no: r
        for r in DeliveryOrder.query.all()
    }

    errors = []
    pending_add = []      # 待新增
    pending_update = []   # 待覆盖（existing对象, data）
    batch_nos = set()     # 本批次内去重

    for row_idx, row in enumerate(rows, start=2):
        try:
            data = _extract_row(row, col_map, DELIVERY_COLUMNS)
            if not data:
                continue

            # 必填校验
            err = _validate_required(data, DELIVERY_REQUIRED, row_idx)
            if err:
                errors.extend(err)
                continue

            # 数值校验
            err = _validate_numeric(data, DELIVERY_NUMERIC, row_idx)
            if err:
                errors.extend(err)
                continue

            # 货物类型默认值
            if not data.get('货物类型') or str(data['货物类型']).strip() == '':
                data['货物类型'] = '普货'

            delivery_no = str(data['交货单号']).strip()

            # 本批次内重复检查（数据质量问题，必须报错）
            if delivery_no in batch_nos:
                errors.append(f'第{row_idx}行：与本次导入中其他行交货单号"{delivery_no}"重复')
                continue
            batch_nos.add(delivery_no)

            existing = existing_map.get(delivery_no)
            if existing:
                if on_duplicate == 'skip':
                    continue  # 静默跳过，不算错误
                else:  # overwrite
                    pending_update.append((existing, data))
                    continue

            # 新增
            order = DeliveryOrder()
            _fill_delivery_order(order, data)
            pending_add.append(order)

        except Exception as e:
            errors.append(f'第{row_idx}行：处理异常 - {str(e)}')

    # 全有或全无（仅数据质量错误）
    if errors:
        return 0, errors

    if pending_add:
        db.session.add_all(pending_add)
    for existing, data in pending_update:
        _fill_delivery_order(existing, data)
    db.session.commit()
    return len(pending_add) + len(pending_update), []


def _normalize_date_str(val):
    """将各种日期值统一转为 YYYYMMDD 字符串（8位数字）。
    支持：datetime/date 对象、"YYYYMMDD"、"YYYY-MM-DD"、"YYYY/MM/DD"、"YYYY-MM-DD HH:MM:SS" 等。
    """
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y%m%d')
    s = str(val).strip()
    if not s:
        return ''
    # 纯 8 位数字已经是 YYYYMMDD
    if len(s) == 8 and s.isdigit():
        return s
    # 去掉 '-' 和 '/' 分隔符再校验
    digits = s[:10].replace('-', '').replace('/', '')
    if len(digits) == 8 and digits.isdigit():
        return digits
    return ''


def _normalize_month_str(val):
    """将各种月份值统一转为 YYYYMM 字符串（6位数字）。
    支持：datetime/date 对象、"YYYYMM"、"YYYY-MM"、"YYYY/MM" 等。
    """
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y%m')
    s = str(val).strip()
    if not s:
        return ''
    # 纯 6 位数字已经是 YYYYMM
    if len(s) == 6 and s.isdigit():
        return s
    # 去掉 '-' 和 '/' 分隔符再校验
    digits = s[:7].replace('-', '').replace('/', '')
    if len(digits) == 6 and digits.isdigit():
        return digits
    return ''


def _fill_delivery_order(order, data):
    order.delivery_no = str(data.get('交货单号', '')).strip()
    order.shipping_point = str(data.get('装运点', '')).strip()
    order.shipping_point_desc = str(data.get('装运点描述', '')).strip()
    order.transport_area = str(data.get('运输区域', '')).strip()
    order.transport_area_desc = str(data.get('运输区域描述', '')).strip()
    order.total_weight = float(data.get('总重量（吨）', 0))
    order.total_volume = float(data.get('总体积（m³）', 0))
    order.consolidated_no = str(data.get('拼车单号', '')).strip()
    order.waybill_no = str(data.get('运单号', '')).strip()
    order.post_date = _normalize_date_str(data.get('发货过账日期', ''))
    order.license_plate = str(data.get('车牌号', '') or '').strip()
    order.sales_org = str(data.get('销售组织', '')).strip()
    order.org_name = str(data.get('组织名称', '')).strip()
    order.carrier_name = str(data.get('承运商名称', '')).strip()
    order.carrier_code = str(data.get('承运商编码', '')).strip()
    order.consignee_name = str(data.get('送达方名称', '')).strip()
    order.consignee_code = str(data.get('送达方编码', '')).strip()
    order.cargo_type = str(data.get('货物类型', '普货')).strip()
    order.transport_mode = str(data.get('运输方式', '') or '').strip()


# ─────────────────────────────────────────
#  零担审批导入
# ─────────────────────────────────────────
def import_ltl_approvals(file_stream, on_duplicate='skip'):
    """
    导入零担审批Excel。
    全有或全无：数据质量错误则整批拒绝。
    on_duplicate: 'skip'=跳过重复  |  'overwrite'=覆盖
    返回 (success_count, errors_list)
    """
    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    col_map = {}
    for i, h in enumerate(header_row):
        if h and str(h).strip() in LTL_COLUMNS:
            col_map[str(h).strip()] = i

    # 一次性加载已有审批记录到 dict
    existing_map = {
        r.delivery_no: r
        for r in LtlApproval.query.all()
    }

    errors = []
    pending_add = []
    pending_update = []
    batch_delivery_nos = set()

    for row_idx, row in enumerate(rows, start=2):
        try:
            data = _extract_row(row, col_map, LTL_COLUMNS)
            if not data:
                continue

            err = _validate_required(data, LTL_REQUIRED, row_idx)
            if err:
                errors.extend(err)
                continue

            ltl_type = str(data['零担类型']).strip()
            if ltl_type not in ('支线', '干线'):
                errors.append(f'第{row_idx}行：零担类型必须为支线/干线')
                continue

            vehicle_seq = str(data['车序号']).strip()
            delivery_no = str(data['交货单号']).strip()
            approval_month = _normalize_month_str(data['零担审批月份'])
            if not approval_month:
                errors.append(f'第{row_idx}行：零担审批月份格式错误，应为YYYYMM')
                continue

            # 本批次内重复（数据质量问题）
            if delivery_no in batch_delivery_nos:
                errors.append(f'第{row_idx}行：与本次导入中其他行交货单号"{delivery_no}"重复')
                continue
            batch_delivery_nos.add(delivery_no)

            ltl_vehicle_no = f'{ltl_type}-{vehicle_seq}'

            existing = existing_map.get(delivery_no)
            if existing:
                if on_duplicate == 'skip':
                    continue
                else:  # overwrite
                    pending_update.append((existing, ltl_type, vehicle_seq, ltl_vehicle_no, approval_month))
                    continue

            rec = LtlApproval(
                ltl_type=ltl_type,
                vehicle_seq=vehicle_seq,
                ltl_vehicle_no=ltl_vehicle_no,
                delivery_no=delivery_no,
                approval_month=approval_month,
            )
            pending_add.append(rec)

        except Exception as e:
            errors.append(f'第{row_idx}行：处理异常 - {str(e)}')

    # 全有或全无
    if errors:
        return 0, errors

    if pending_add:
        db.session.add_all(pending_add)
    for rec, lt, vs, lvn, am in pending_update:
        rec.ltl_type = lt
        rec.vehicle_seq = vs
        rec.ltl_vehicle_no = lvn
        rec.approval_month = am
    db.session.commit()
    return len(pending_add) + len(pending_update), []


# ─────────────────────────────────────────
#  合同导入
# ─────────────────────────────────────────
def import_contracts(file_stream):
    """
    导入合同列表Excel。
    全有或全无：任一行校验失败则整批拒绝。
    返回 (success_count, errors_list)
    """
    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    col_map = {}
    for i, h in enumerate(header_row):
        if h and str(h).strip() in CONTRACT_COLUMNS:
            col_map[str(h).strip()] = i

    # 一次性加载已有合同编码到 Set
    existing_codes = set(
        r[0] for r in Contract.query.with_entities(Contract.contract_code).all()
    )

    errors = []
    pending = []
    batch_codes = set()

    for row_idx, row in enumerate(rows, start=2):
        try:
            data = _extract_row(row, col_map, CONTRACT_COLUMNS)
            if not data:
                continue

            err = _validate_required(data, CONTRACT_REQUIRED, row_idx)
            if err:
                errors.extend(err)
                continue

            contract_code = str(data['合同编码']).strip()

            # 数据库已有
            if contract_code in existing_codes:
                errors.append(f'第{row_idx}行：合同编码"{contract_code}"已存在')
                continue

            # 本批次内重复
            if contract_code in batch_codes:
                errors.append(f'第{row_idx}行：与本次导入中其他行合同编码"{contract_code}"重复')
                continue
            batch_codes.add(contract_code)

            # 日期格式转换
            start_date = _normalize_date(data.get('合同有效期起始'))
            end_date = _normalize_date(data.get('合同有效期截止'))
            if not start_date or not end_date:
                errors.append(f'第{row_idx}行：日期格式错误，应为YYYYMMDD')
                continue

            c = Contract(
                contract_code=contract_code,
                contract_name=str(data['合同名称']).strip(),
                carrier_name=str(data['承运商名称']).strip(),
                carrier_code=str(data['承运商编码']).strip(),
                start_date=start_date,
                end_date=end_date,
                status='草稿',  # 导入时默认草稿
            )
            pending.append(c)

        except Exception as e:
            errors.append(f'第{row_idx}行：处理异常 - {str(e)}')

    # 全有或全无
    if errors:
        return 0, errors

    db.session.add_all(pending)
    db.session.commit()
    return len(pending), []


# ─────────────────────────────────────────
#  费率导入
# ─────────────────────────────────────────
def import_rates(contract_id, file_stream):
    """
    向指定合同导入费率Excel。
    全有或全无：任一行校验失败则整批拒绝。
    校验：必填、数值、价格连续性、重复性。
    返回 (success_count, errors_list)
    """
    contract = Contract.query.get(contract_id)
    if not contract:
        return 0, ['合同不存在']

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    col_map = {}
    for i, h in enumerate(header_row):
        if h and str(h).strip() in RATE_COLUMNS:
            col_map[str(h).strip()] = i

    # 一次性加载该合同已有费率组合到 Set
    existing_rate_keys = set(
        (r.route_code, r.cargo_type, r.transport_mode)
        for r in ContractRate.query.filter_by(contract_id=contract_id)
        .with_entities(ContractRate.route_code, ContractRate.cargo_type, ContractRate.transport_mode)
        .all()
    )

    errors = []
    pending_rates = []       # 校验通过的待入库记录
    seen_keys = set()        # 本批次内去重

    for row_idx, row in enumerate(rows, start=2):
        try:
            data = _extract_row(row, col_map, RATE_COLUMNS)
            if not data:
                continue

            err = _validate_required(data, RATE_REQUIRED, row_idx)
            if err:
                errors.extend(err)
                continue

            err = _validate_numeric(data, RATE_NUMERIC, row_idx)
            if err:
                errors.extend(err)
                continue

            # 价格连续性校验
            prices = [
                float(data.get('含增值税运输价格1', 0) or 0),
                float(data.get('含增值税运输价格2', 0) or 0),
                float(data.get('含增值税运输价格3', 0) or 0),
                float(data.get('含增值税运输价格4', 0) or 0),
            ]
            tier_count = 0
            for p in prices:
                if p < 0:
                    errors.append(f'第{row_idx}行：含增值税运输价格不能为负数')
                    break
            else:
                for p in prices:
                    if p > 0:
                        tier_count += 1
                    else:
                        break
                continuous = True
                for i_p in range(tier_count):
                    if prices[i_p] == 0:
                        continuous = False
                        break

                if not continuous:
                    errors.append(f'第{row_idx}行：坎级价格不连续，价格必须从价格1开始连续非零')
                    continue

            # 税率校验
            tax_rate = float(data.get('增值税税率', 0) or 0)
            if tax_rate > 1:
                errors.append(f'第{row_idx}行：增值税税率超过100%')
                continue

            # 自动生成字段
            sp = str(data.get('装运点', '')).strip()
            ta = str(data.get('运输区域', '')).strip()
            route_code = f'{sp}&{ta}'

            terminal_fee = float(data.get('含增值税末端费用', 0) or 0)
            has_terminal = '是' if terminal_fee > 0 else '否'

            cargo_type = str(data.get('货物类型', '') or '').strip()
            if not cargo_type:
                cargo_type = '普货'

            transport_mode = str(data.get('运输方式', '') or '').strip()

            # 重复性校验：数据库已有
            rate_key = (route_code, cargo_type, transport_mode)
            if rate_key in existing_rate_keys:
                errors.append(f'第{row_idx}行：线路{route_code} + 货物类型{cargo_type} + 运输方式{transport_mode} 已存在费率')
                continue

            # 重复性校验：本批次内重复
            if rate_key in seen_keys:
                errors.append(f'第{row_idx}行：与本次导入中其他行重复（线路{route_code} + 货物类型{cargo_type} + 运输方式{transport_mode}）')
                continue
            seen_keys.add(rate_key)

            rate = ContractRate(
                contract_id=contract_id,
                route_code=route_code,
                tier_count=tier_count,
                shipping_point=sp,
                shipping_point_desc=str(data.get('装运点描述', '')).strip(),
                transport_area=ta,
                transport_area_desc=str(data.get('运输区域描述', '')).strip(),
                province=str(data.get('省/直辖市', '')).strip(),
                city=str(data.get('地级市', '')).strip(),
                district=str(data.get('县/区', '')).strip(),
                mileage=float(data.get('里程(KM)', 0) or 0),
                route_type=str(data.get('线路类型', '')).strip(),
                has_terminal=has_terminal,
                tax_rate=tax_rate,
                price1=prices[0],
                price2=prices[1],
                price3=prices[2],
                price4=prices[3],
                terminal_fee=terminal_fee,
                cargo_type=cargo_type,
                transport_mode=transport_mode,
            )
            pending_rates.append(rate)

        except Exception as e:
            errors.append(f'第{row_idx}行：处理异常 - {str(e)}')

    # 全有或全无：有错误则全部拒绝
    if errors:
        return 0, errors

    db.session.add_all(pending_rates)
    db.session.commit()
    return len(pending_rates), []


# ─────────────────────────────────────────
#  试算结果导出
# ─────────────────────────────────────────
def export_trial_results(results, columns):
    """
    将试算结果导出为 Excel。
    results: list of dict
    columns: list of column names to export
    返回 BytesIO 对象
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '试算结果'

    # 表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────
#  交货单导出（模板）
# ─────────────────────────────────────────
def create_delivery_template():
    """生成交货单导入模板"""
    return _create_template(DELIVERY_COLUMNS)


def create_ltl_template():
    """生成零担审批导入模板"""
    return _create_template(LTL_COLUMNS)


def create_contract_template():
    """生成合同导入模板"""
    return _create_template(CONTRACT_COLUMNS)


def create_rate_template():
    """生成费率导入模板"""
    return _create_template(RATE_COLUMNS)


def _create_template(columns):
    wb = Workbook()
    ws = wb.active
    ws.title = '导入模板'
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) * 2, 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────
#  工具函数
# ─────────────────────────────────────────
def _extract_row(row, col_map, expected_columns):
    """根据列映射从行数据提取字典"""
    data = {}
    for col_name in expected_columns:
        if col_name in col_map:
            idx = col_map[col_name]
            val = row[idx] if idx < len(row) else None
            data[col_name] = val
    return data


def _validate_required(data, required_fields, row_idx):
    errors = []
    for field in required_fields:
        val = data.get(field)
        if val is None or str(val).strip() == '':
            errors.append(f'第{row_idx}行，字段{field}不能为空')
    return errors


def _validate_numeric(data, numeric_fields, row_idx):
    errors = []
    for field in numeric_fields:
        val = data.get(field)
        if val is not None and str(val).strip() != '':
            try:
                float(val)
            except (ValueError, TypeError):
                errors.append(f'第{row_idx}行，字段{field}格式不正确（应为数值）')
    return errors


def _normalize_date(val):
    """
    将日期值统一转为 YYYYMMDD 字符串。
    支持：datetime对象、int(20240101)、str('20240101'/'2024-01-01')
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y%m%d')
    s = str(val).strip()
    s = s.replace('-', '').replace('/', '')
    if len(s) == 8 and s.isdigit():
        return s
    return None
